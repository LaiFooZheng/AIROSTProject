from openpyxl import load_workbook
import pyotp
import otp

print("Loading...")
database = load_workbook('userdata.xlsx')
user_data = database["Datas"]

user_count = user_data.max_column
nam = user_data['1']
acc = user_data['2']
pwd = user_data['3']
balance = user_data['4']
email = user_data['5']
hp_num = user_data['6']
secret_key = user_data['7']
face = user_data['8':'135']
print("Database loaded with", user_count, "user(s).")


def reload():
    global nam, acc, pwd, balance, email, hp_num, secret_key, face
    nam = user_data['1']
    acc = user_data['2']
    pwd = user_data['3']
    balance = user_data['4']
    email = user_data['5']
    hp_num = user_data['6']
    secret_key = user_data['7']
    face = user_data['8':'135']
    print("Database reloaded with", user_count, "user(s).")


def save_2():
    database.save('userdata.xlsx')


def backup():
    database.save('userdata.bak')


def registration(user_id, user_name, password, user_email, user_hp_num, face_id):
    user_data.cell(row=2, column=user_count+1, value=user_id)
    reload()
    nam[-1].value = user_name
    pwd[-1].value = password
    balance[-1].value = 0
    email[-1].value = user_email
    hp_num[-1].value = user_hp_num
    secret_key[-1].value = pyotp.random_base32()
    otp.otp_generate(secret_key[-1].value, user_email)
    for x in range(len(face_id)):
        face[x][-1].value = face_id[x]
    save_2()
    print("Account created!")
